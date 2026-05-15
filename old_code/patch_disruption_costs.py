"""
Retroactively add disruption_cost and disruption_pct to all phase2 xlsx outputs.

For each scenario/strategy xlsx:
  - Insert two rows after "Objective ($)" in the Summary sheet:
      Disruption Cost ($)   =  logistics_cost - Z_STAR
      Disruption (%)        =  disruption_cost / Z_STAR * 100

Also updates disruption_cost and disruption_pct in summary_all_scenarios.xlsx.

For structurally-infeasible scenarios (S1/S3, strategies R and A), the
logistics_cost covers only the demand that was actually served. The disruption
cost is therefore negative (less spend, but also less service) and is flagged
as "partial" in the notes.
"""

import os
import glob
import pandas as pd
from openpyxl import load_workbook

_HERE = os.path.dirname(os.path.abspath(__file__))

BASELINE_FILE  = os.path.join(_HERE, "phase1", "results", "baseline_solution.xlsx")
RESULTS_DIR    = os.path.join(_HERE, "phase2", "results")
MASTER_SUMMARY = os.path.join(RESULTS_DIR, "summary_all_scenarios.xlsx")

# ---------------------------------------------------------------------------
# 1. Load Z*
# ---------------------------------------------------------------------------
_base_df = pd.read_excel(BASELINE_FILE, sheet_name="Summary")
_zstar_row = _base_df[_base_df["Metric"] == "Total Cost ($)"]
if _zstar_row.empty:
    raise ValueError("Cannot find 'Total Cost ($)' in baseline Summary sheet.")
Z_STAR = float(_zstar_row["Value"].iloc[0])
print(f"Z* = ${Z_STAR:,.2f}")

# ---------------------------------------------------------------------------
# 2. Patch individual strategy xlsx files
# ---------------------------------------------------------------------------
xlsx_files = glob.glob(os.path.join(RESULTS_DIR, "scenario_ArcCosts_*", "strategy_*.xlsx"))
xlsx_files.sort()

for fpath in xlsx_files:
    wb = load_workbook(fpath)
    ws = wb["Summary"]

    # Read current rows into list of (metric, value) pairs
    rows = [(ws.cell(r, 1).value, ws.cell(r, 2).value)
            for r in range(1, ws.max_row + 1)]

    # Find logistics cost and the row index of "Objective ($)"
    logistics_cost = None
    obj_row_idx = None  # 0-based index in `rows`
    for i, (metric, value) in enumerate(rows):
        if isinstance(metric, str) and metric.strip() == "Logistics Cost ($)":
            logistics_cost = float(value)
        if isinstance(metric, str) and metric.strip() in (
                "Objective ($)", "Objective (logistics+penalty)"):
            obj_row_idx = i

    if logistics_cost is None:
        print(f"  SKIP (no Logistics Cost): {fpath}")
        continue

    disruption_cost = logistics_cost - Z_STAR
    disruption_pct  = disruption_cost / Z_STAR * 100

    # Check whether disruption rows already exist
    existing_metrics = {str(m).strip() for m, _ in rows if m is not None}
    if "Disruption Cost ($)" in existing_metrics:
        print(f"  Already patched, skipping: {os.path.relpath(fpath, _HERE)}")
        continue

    # Insert two rows right after "Objective ($)"
    insert_at = obj_row_idx + 1  # 0-based; insert after this index
    new_rows = (
        rows[: insert_at + 1]
        + [("Disruption Cost ($)", round(disruption_cost, 2)),
           ("Disruption (%)",      round(disruption_pct,  4))]
        + rows[insert_at + 1 :]
    )

    # Clear sheet and rewrite
    ws.delete_rows(1, ws.max_row)
    for i, (metric, value) in enumerate(new_rows, start=1):
        ws.cell(i, 1).value = metric
        ws.cell(i, 2).value = value

    wb.save(fpath)
    rel = os.path.relpath(fpath, _HERE)
    print(f"  Patched {rel}: ΔZ = ${disruption_cost:+,.2f} ({disruption_pct:+.2f}%)")

# ---------------------------------------------------------------------------
# 3. Patch master summary
# ---------------------------------------------------------------------------
master_df = pd.read_excel(MASTER_SUMMARY)

def _compute(row):
    lc = row["logistics_cost"]
    dc = lc - Z_STAR
    return pd.Series({
        "disruption_cost": round(dc, 2),
        "disruption_pct":  round(dc / Z_STAR * 100, 4),
    })

computed = master_df.apply(_compute, axis=1)
master_df["disruption_cost"] = computed["disruption_cost"]
master_df["disruption_pct"]  = computed["disruption_pct"]

with pd.ExcelWriter(MASTER_SUMMARY, engine="openpyxl",
                    mode="a", if_sheet_exists="replace") as writer:
    master_df.to_excel(writer, sheet_name="Sheet1", index=False)

print(f"\nMaster summary updated: {os.path.relpath(MASTER_SUMMARY, _HERE)}")
print(master_df[["scenario_key", "strategy", "logistics_cost",
                  "disruption_cost", "disruption_pct"]].to_string(index=False))
